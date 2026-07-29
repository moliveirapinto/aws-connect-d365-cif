"""Build the live-transcription cost model workbook (incremental cost only, telephony excluded)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = Font(name="Arial", color="0000FF")          # inputs
BLACK = Font(name="Arial", color="000000")         # formulas
BLACKB = Font(name="Arial", color="000000", bold=True)
WHITEB = Font(name="Arial", color="FFFFFF", bold=True)
GREEN = Font(name="Arial", color="008000")         # cross-sheet links
TITLE = Font(name="Arial", size=14, bold=True, color="1F3864")
NOTE = Font(name="Arial", size=9, italic=True, color="808080")

YELLOW = PatternFill("solid", fgColor="FFFF00")
HDR = PatternFill("solid", fgColor="1F3864")
LTGREY = PatternFill("solid", fgColor="F2F2F2")
ACCENT = PatternFill("solid", fgColor="D9E1F2")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

USD = '$#,##0.00;($#,##0.00);"-"'
USD0 = '$#,##0;($#,##0);"-"'
USD4 = '$#,##0.0000;($#,##0.0000);"-"'
NUM = '#,##0;(#,##0);"-"'

wb = Workbook()

# ============================== SHEET 1: MODEL ==============================
ws = wb.active
ws.title = "Cost Model"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 44
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 50

def title(cell, text):
    ws[cell] = text
    ws[cell].font = TITLE

def hdr(row, label):
    c = ws.cell(row=row, column=2, value=label)
    c.font = WHITEB
    c.fill = HDR
    for col in range(2, 5):
        ws.cell(row=row, column=col).fill = HDR
    ws.cell(row=row, column=3).font = WHITEB
    ws.cell(row=row, column=4).font = WHITEB

title("B2", "AWS Connect → Azure Speech → D365  |  Live Transcription Cost Model")
ws["B3"] = "Incremental cost of the transcription layer only. EXCLUDES telephony (existing cost) and D365 Copilot licensing (separate)."
ws["B3"].font = NOTE

# ---- INPUTS ----
hdr(5, "INPUTS  (edit the blue cells)")
ws["C5"] = "Value"; ws["C5"].font = WHITEB; ws["C5"].alignment = Alignment(horizontal="center")
ws["D5"] = "Unit"; ws["D5"].font = WHITEB; ws["D5"].alignment = Alignment(horizontal="center")

inputs = [
    ("Call minutes per month",                 11000,  "min",       NUM),
    ("Channels transcribed (1 = caller only, 2 = both)", 2, "ch",   '#,##0'),
    ("Azure Speech price",                       1.00,  "$/audio-hr", USD),
    ("Direct Line segments per call-minute",      2.0,  "msg/min",   '#,##0.0'),
    ("Direct Line price",                         0.50, "$/1000 msg", USD),
    ("Kinesis Video Streams price",              0.0085,"$/GB",      USD4),
    ("AWS→Azure egress price",                   0.09,  "$/GB",      USD),
    ("Audio volume per call-hour, per channel",  0.12,  "GB",        '#,##0.00'),
    ("Container Apps (consumer + ingestor), fixed", 65, "$/month",   USD0),
]
r = 6
input_rows = {}
for label, val, unit, fmt in inputs:
    ws.cell(row=r, column=2, value=label).font = BLACK
    c = ws.cell(row=r, column=3, value=val)
    c.font = BLUE; c.fill = YELLOW; c.number_format = fmt
    c.alignment = Alignment(horizontal="center"); c.border = BORDER
    ws.cell(row=r, column=4, value=unit).font = NOTE
    ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
    input_rows[label] = r
    r += 1

MIN   = f"$C${input_rows['Call minutes per month']}"
CH    = f"$C${input_rows['Channels transcribed (1 = caller only, 2 = both)']}"
SPEECH= f"$C${input_rows['Azure Speech price']}"
SEG   = f"$C${input_rows['Direct Line segments per call-minute']}"
DLP   = f"$C${input_rows['Direct Line price']}"
KVSP  = f"$C${input_rows['Kinesis Video Streams price']}"
EGP   = f"$C${input_rows['AWS→Azure egress price']}"
GBHR  = f"$C${input_rows['Audio volume per call-hour, per channel']}"
CAPPS = f"$C${input_rows['Container Apps (consumer + ingestor), fixed']}"

# ---- CALCULATIONS ----
r += 1
hdr(r, "MONTHLY COST BREAKDOWN")
ws.cell(row=r, column=3, value="Cost").font = WHITEB
ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
ws.cell(row=r, column=4, value="% of total").font = WHITEB
ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
calc_start = r + 1

def calc(row, label, formula, note=""):
    ws.cell(row=row, column=2, value=label).font = BLACK
    c = ws.cell(row=row, column=3, value=formula)
    c.font = BLACK; c.number_format = USD; c.border = BORDER
    if note:
        ws.cell(row=row, column=5, value=note).font = NOTE

r = calc_start
calc(r, "Azure Speech (real-time STT)",
     f"=({MIN}/60)*{CH}*{SPEECH}",
     "= (minutes / 60) × channels × $/audio-hour")
speech_row = r; r += 1
calc(r, "Bot Direct Line messages",
     f"=({MIN}*{SEG}/1000)*{DLP}",
     "= (minutes × segments / 1000) × $/1000 msg")
dl_row = r; r += 1
calc(r, "Kinesis Video Streams (ingest + consume)",
     f"=({MIN}/60)*{GBHR}*{CH}*2*{KVSP}",
     "= call-hours × GB/hr/ch × channels × 2 × $/GB")
kvs_row = r; r += 1
calc(r, "AWS→Azure egress",
     f"=({MIN}/60)*{GBHR}*{CH}*{EGP}",
     "= call-hours × GB/hr/ch × channels × $/GB")
eg_row = r; r += 1
calc(r, "Container Apps (fixed)",
     f"={CAPPS}",
     "flat, regardless of call volume")
capps_row = r; r += 1

# total
ws.cell(row=r, column=2, value="TOTAL MONTHLY COST").font = BLACKB
tc = ws.cell(row=r, column=3, value=f"=SUM(C{speech_row}:C{capps_row})")
tc.font = BLACKB; tc.number_format = USD; tc.fill = ACCENT; tc.border = BORDER
total_row = r
for col in range(2, 5):
    ws.cell(row=r, column=col).fill = ACCENT
r += 1

# per-minute
ws.cell(row=r, column=2, value="Cost per call-minute (all-in)").font = BLACKB
pm = ws.cell(row=r, column=3, value=f"=C{total_row}/{MIN}")
pm.font = BLACKB; pm.number_format = USD4; pm.border = BORDER
r += 1
ws.cell(row=r, column=2, value="Variable cost per call-minute (excl. fixed)").font = BLACK
pv = ws.cell(row=r, column=3, value=f"=(C{total_row}-C{capps_row})/{MIN}")
pv.font = BLACK; pv.number_format = USD4; pv.border = BORDER
r += 1

# % of total column
for rr in [speech_row, dl_row, kvs_row, eg_row, capps_row]:
    c = ws.cell(row=rr, column=4, value=f"=C{rr}/$C${total_row}")
    c.font = BLACK; c.number_format = "0.0%"; c.alignment = Alignment(horizontal="center")

# ============================== SHEET 2: SCENARIOS ==============================
ws2 = wb.create_sheet("Scenarios")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 2
ws2.column_dimensions["B"].width = 34
for col in "CDE":
    ws2.column_dimensions[col].width = 16

ws2["B2"] = "Monthly Cost by Volume"
ws2["B2"].font = TITLE
ws2["B3"] = "Uses the unit prices from the 'Cost Model' sheet. Only call minutes differ per column."
ws2["B3"].font = NOTE

# scenario header
hrow = 5
labels = ["Assumption / Scenario", "Small (pilot)", "Medium", "Large"]
for i, lab in enumerate(labels):
    c = ws2.cell(row=hrow, column=2+i, value=lab)
    c.font = WHITEB; c.fill = HDR
    c.alignment = Alignment(horizontal="center" if i else "left")

# minutes input row (blue)
mrow = hrow + 1
ws2.cell(row=mrow, column=2, value="Call minutes / month").font = BLACK
for i, val in enumerate([2000, 11000, 100000]):
    c = ws2.cell(row=mrow, column=3+i, value=val)
    c.font = BLUE; c.fill = YELLOW; c.number_format = NUM
    c.alignment = Alignment(horizontal="center"); c.border = BORDER

# pull unit prices from model sheet (green cross-sheet)
def mref(cellref):
    return f"'Cost Model'!{cellref}"

CHs   = mref(CH); SPs = mref(SPEECH); SEGs = mref(SEG); DLPs = mref(DLP)
KVSPs = mref(KVSP); EGPs = mref(EGP); GBHRs = mref(GBHR); CAPPSs = mref(CAPPS)

rows2 = [
    ("Azure Speech", lambda m: f"=({m}/60)*{CHs}*{SPs}"),
    ("Direct Line", lambda m: f"=({m}*{SEGs}/1000)*{DLPs}"),
    ("KVS + egress", lambda m: f"=({m}/60)*{GBHRs}*{CHs}*2*{KVSPs}+({m}/60)*{GBHRs}*{CHs}*{EGPs}"),
    ("Container Apps (fixed)", lambda m: f"={CAPPSs}"),
]
rr = mrow + 1
first_calc = rr
for label, fx in rows2:
    ws2.cell(row=rr, column=2, value=label).font = BLACK
    for i in range(3):
        mcell = f"{get_column_letter(3+i)}{mrow}"
        c = ws2.cell(row=rr, column=3+i, value=fx(mcell))
        c.font = GREEN if label != "Container Apps (fixed)" else GREEN
        c.number_format = USD0; c.border = BORDER
    rr += 1
last_calc = rr - 1

# total
ws2.cell(row=rr, column=2, value="TOTAL / MONTH").font = BLACKB
for i in range(3):
    col = get_column_letter(3+i)
    c = ws2.cell(row=rr, column=3+i, value=f"=SUM({col}{first_calc}:{col}{last_calc})")
    c.font = BLACKB; c.number_format = USD0; c.fill = ACCENT; c.border = BORDER
trow = rr
rr += 1
# annual
ws2.cell(row=rr, column=2, value="TOTAL / YEAR").font = BLACKB
for i in range(3):
    col = get_column_letter(3+i)
    c = ws2.cell(row=rr, column=3+i, value=f"={col}{trow}*12")
    c.font = BLACKB; c.number_format = USD0; c.border = BORDER
rr += 1
# per minute
ws2.cell(row=rr, column=2, value="Cost per call-minute").font = BLACK
for i in range(3):
    col = get_column_letter(3+i)
    c = ws2.cell(row=rr, column=3+i, value=f"={col}{trow}/{col}{mrow}")
    c.font = BLACK; c.number_format = USD4; c.border = BORDER

rr += 2
ws2.cell(row=rr, column=2, value="Legend:").font = NOTE
ws2.cell(row=rr+1, column=2, value="Blue = editable input   •   Green = pulled from Cost Model sheet   •   Black = formula").font = NOTE

# ============================== SHEET 3: NOTES ==============================
ws3 = wb.create_sheet("Assumptions & Sources")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 2
ws3.column_dimensions["B"].width = 100
ws3["B2"] = "Assumptions, Method & Caveats"
ws3["B2"].font = TITLE
notes = [
    "",
    "SCOPE — this model prices ONLY the net-new transcription pipeline that was built:",
    "   AWS Kinesis Video Streams → Azure Container Apps (KVS consumer) → Azure AI Speech (real-time STT) → Azure Bot Direct Line → D365.",
    "",
    "EXCLUDED on purpose:",
    "   • Amazon Connect telephony / per-minute call charges — this is an existing cost the customer already pays.",
    "   • D365 Copilot for Service licensing (agent assist, summaries) — a separate per-seat / capacity SKU.",
    "",
    "KEY DRIVER:",
    "   • Azure Speech is ~90–95% of the incremental cost. It is billed per audio-hour, PER channel.",
    "   • Transcribing both caller and agent = 2 channels = 2× Speech. Set 'Channels' to 1 to transcribe caller only (~half the Speech cost).",
    "",
    "UNIT-PRICE SOURCES (US, list price — confirm for your region/commitment):",
    "   • Azure AI Speech, standard real-time speech-to-text: $1.00 / audio-hour. Commitment tiers can lower this to ~$0.60–0.80/hr at volume.",
    "   • Azure Bot Service Direct Line (S1): $0.50 per 1,000 messages. Each transcript segment ≈ 1 message (~2/min assumed).",
    "   • Amazon Kinesis Video Streams: ~$0.0085/GB ingested + ~$0.0085/GB consumed.",
    "   • AWS→Azure data egress: ~$0.09/GB.",
    "   • Azure Container Apps: ~$30–35/app/month at 0.5 vCPU / 1 GB, always-on (min 1 replica); two apps ≈ $65/mo after free grant.",
    "",
    "AUDIO MATH:",
    "   • 8 kHz / 16-bit / mono ≈ 0.96 MB per minute per track ≈ 0.12 GB per call-hour per channel.",
    "",
    "All prices are list and approximate; use for planning/estimation, not billing. Verify on the Azure and AWS pricing calculators for your region.",
]
rn = 3
for line in notes:
    c = ws3.cell(row=rn, column=2, value=line)
    c.font = NOTE if line.startswith("   ") or line == "" else Font(name="Arial", bold=line.endswith(":") or line.startswith(("SCOPE","EXCLUDED","KEY","UNIT","AUDIO")), color="000000")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    rn += 1

try:
    wb.calculation.fullCalcOnLoad = True
except Exception:
    pass

out = r"c:\Users\maoliveira\dev\aws-connect-d365-cif\docs\LiveTranscription-CostModel.xlsx"
wb.save(out)
print("saved", out)

# ---- independent sanity check of the default-input math (no Excel needed) ----
minutes, ch, speech, seg, dlp, kvsp, egp, gbhr, capps = 11000, 2, 1.00, 2, 0.5, 0.0085, 0.09, 0.12, 65
sp = (minutes/60)*ch*speech
dl = (minutes*seg/1000)*dlp
kvs = (minutes/60)*gbhr*ch*2*kvsp
eg = (minutes/60)*gbhr*ch*egp
tot = sp+dl+kvs+eg+capps
print(f"CHECK  speech={sp:.2f} dl={dl:.2f} kvs={kvs:.2f} egress={eg:.2f} fixed={capps} "
      f"TOTAL/mo={tot:.2f}  $/min={tot/minutes:.4f}  speech%%={sp/tot*100:.1f}")
